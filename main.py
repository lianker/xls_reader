import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


spath = '{}/sheets/Modelo.xlsx'.format(os.getcwd())


def opening_sheets(path_to_sheet):
    print('=========')
    wb = openpyxl.load_workbook(path_to_sheet)
    print(type(wb))
    print('=========')
    sheet_names = wb.sheetnames
    print(sheet_names)
    print('=========')
    exemplo = wb['exemplo']
    print(type(exemplo))
    print('=========')
    work = wb.active
    print(work)


def reading_cells(path_to_sheet):
    wb = openpyxl.load_workbook(path_to_sheet)
    sheet = wb['exemplo']
    print(sheet['A1'])
    print(type(sheet['A1']))
    print(sheet['C2'].value)
    # colunas agora é numero nao letra que nem no livro
    c = sheet['C1']
    print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)

    # Usando o método cell()
    b = sheet.cell(row=1, column=2)
    print('Row ' + str(b.row) + ', Column ' + str(b.column) + ' is ' + b.value)

    # Laço
    print(range(1, 8, 2))
    for i in range(1, 8, 2):
        print(i, sheet.cell(row=i, column=2).value)


def sheet_size(path_to_sheet):
    wb = openpyxl.load_workbook(path_to_sheet)
    sheet = wb['exemplo']
    # highest_row = sheet.get_highest_row() antigo
    row_count = sheet.max_row
    column_count = sheet.max_column
    print(row_count)
    print(column_count)


def column_names():
    print('Column letter: ' + get_column_letter(1))
    print('index: ' + str(column_index_from_string('a')))


def main():
    column_names()


if __name__ == "__main__":
    main()
